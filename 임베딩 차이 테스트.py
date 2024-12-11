import os
import openpyxl
from openpyxl.styles import PatternFill
from langchain_openai import OpenAIEmbeddings

# 예제 텍스트
text1 = '''대출의 기간연장은 만기일로부터 1개월 전부터 신청이 가능합니다.개인대출의 경우 자동연장(녹취에 의한 전화 연장)이 가능하며, 자동연장 대상인 경우 고객센터 자동연장팀에서 대출만기 7일~10일 전에 고객님께 연락을 드려 연장의사 확인 후 진행해 드리고 있습니다.단, 자동연장 제외상품(예: 기업대출, 외화대출, 전세자금대출 등), 연락불가, 녹취거절, 영업점 연장 희망 고객 등은 유선연장이 불가하므로 필요한 서류를 지참하여 영업점 방문하여 신청해야 하며, 인터넷/우리WON뱅킹에서 실행한 대출(군인공제회 회원대출 제외)은 인터넷/우리WON뱅킹에서 연장 가능합니다.유선연장, 인터넷/우리WON뱅킹 연장을 제외한 일반대출 연장은 대출만기 1개월 이내에 대출관리지점 또는 고객센터로 문의해 주시면 연장관련 안내를 받으실 수 있습니다.'''
text2 = '''대출의 기간연장은 만기일로부터 1개월 전부터 신청이 가능합니다.
개인대출의 경우 자동연장(녹취에 의한 전화 연장)이 가능하며, 자동연장 대상인 경우 고객센터 자동연장팀에서 대출만기 7일~10일 전에 고객님께 연락을 드려 연장의사 확인 후 진행해 드리고 있습니다.
단, 자동연장 제외상품(예: 기업대출, 외화대출, 전세자금대출 등), 연락불가, 녹취거절, 영업점 연장 희망 고객 등은 유선연장이 불가하므로 필요한 서류를 지참하여 영업점 방문하여 신청해야 하며, 인터넷/우리WON뱅킹에서 실행한 대출(군인공제회 회원대출 제외)은 인터넷/우리WON뱅킹에서 연장 가능합니다.
유선연장, 인터넷/우리WON뱅킹 연장을 제외한 일반대출 연장은 대출만기 1개월 이내에 대출관리지점 또는 고객센터로 문의해 주시면 연장관련 안내를 받으실 수 있습니다.'''


# 임베딩 모델 초기화
embedding_model = OpenAIEmbeddings(model="text-embedding-3-small")

# 임베딩 계산
embedding1 = embedding_model.embed_query(text1)
embedding2 = embedding_model.embed_query(text2)

# 임베딩 길이 확인 (기본 1536 차원)
dim = len(embedding1)
assert len(embedding2) == dim, "두 임베딩의 길이가 다릅니다."

# 엑셀 워크북 및 워크시트 생성
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Embedding Differences"

# 헤더 작성
ws.cell(row=1, column=1, value="Index")
ws.cell(row=1, column=2, value="Text1_Embedding")
ws.cell(row=1, column=3, value="Text2_Embedding")

# 차원별 값 기록
for i in range(dim):
    row_idx = i + 2  # 2행부터 값 기록 (1행은 헤더)

    val1 = embedding1[i]
    val2 = embedding2[i]

    ws.cell(row=row_idx, column=1, value=i)
    ws.cell(row=row_idx, column=2, value=val1)
    ws.cell(row=row_idx, column=3, value=val2)

    # 값 비교 후 색상 설정
    if val1 == val2:
        # 값이 같을 경우 파란색
        fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
        ws.cell(row=row_idx, column=2).fill = fill
        ws.cell(row=row_idx, column=3).fill = fill
    else:
        # 값이 다를 경우 빨간색
        fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        ws.cell(row=row_idx, column=2).fill = fill
        ws.cell(row=row_idx, column=3).fill = fill

# 엑셀 파일 저장
filename = "embedding_difference.xlsx"
wb.save(filename)

# 절대 경로 출력
abs_path = os.path.abspath(filename)
print(abs_path)