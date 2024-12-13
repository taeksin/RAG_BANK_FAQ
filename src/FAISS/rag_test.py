import os
import time
import pandas as pd
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from concurrent.futures import ProcessPoolExecutor
from langchain import hub
from langchain.prompts import PromptTemplate
from langchain_core.output_parsers import StrOutputParser
from langchain_core.runnables import RunnablePassthrough
from langchain_openai import OpenAIEmbeddings, ChatOpenAI
from langchain_community.vectorstores import FAISS
from dotenv import load_dotenv

# 설정
FAISS_FOLDER = "data/faiss_index_clean"
PROMPT_TEXT = "src/FAISS/prompt.txt"
EXCEL_FILE = "woori_faq_전처리 전후 비교.xlsx"

# .env 파일 로드
load_dotenv()

# OpenAI API 설정
api_key = os.getenv("OPENAI_API_KEY")
os.environ["OPENAI_API_KEY"] = api_key

# 임베딩 모델 생성
embedding_model = OpenAIEmbeddings(model="text-embedding-3-small")

# FAISS 저장소 경로 설정
base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
faiss_path = os.path.join(base_dir, f"{FAISS_FOLDER}")

# Excel 파일 경로
excel_file_path = os.path.join(base_dir, f"data/data_source/{EXCEL_FILE}")

# 벡터 저장소 로드
vectorstore = FAISS.load_local(faiss_path, embeddings=embedding_model, allow_dangerous_deserialization=True)
retriever = vectorstore.as_retriever(search_type='mmr', search_kwargs={'k': 5, 'fetch_k': 10, 'lambda_mult': 0.9})

# 프롬프트 파일 경로 설정
prompt_file_path = os.path.join(base_dir, f"{PROMPT_TEXT}")

# 텍스트 파일을 읽어와 프롬프트 텍스트로 저장
def load_prompt_from_file(file_path):
    with open(file_path, "r", encoding="utf-8") as file:
        return file.read()

# 파일에서 프롬프트 텍스트 읽기
prompt_text = load_prompt_from_file(prompt_file_path)

# PromptTemplate 설정
prompt = PromptTemplate(
    input_variables=["question", "context"],  # 필요한 입력 변수 설정
    template=prompt_text  # 텍스트 파일에서 읽어온 프롬프트 텍스트
)

# LLM 모델 설정
llm = ChatOpenAI(model_name="gpt-4o", temperature=0.5)

# RAG 체인 설정
rag_chain = (
    {"context": retriever, "question": RunnablePassthrough()}  # 질의 및 컨텍스트 설정
    | prompt  # 읽어온 텍스트 파일을 템플릿으로 사용
    | llm  # OpenAI의 GPT 모델을 사용하여 응답 생성
    | StrOutputParser()  # 문자열로 출력
)



# 질문 처리 함수
def process_question(args):
    index, question = args
    if not question.strip():
        return index, ""

    try:
        retrieved_documents = retriever.invoke(question)
        response = rag_chain.invoke(question)
        return index, response
    except Exception as e:
        print(f"질문 처리 중 오류 발생: {e}")
        return index, "오류 발생"

# 기존 엑셀 파일에 데이터 추가 함수
def save_to_existing_excel(file_path, df):
    try:
        # 기존 엑셀 파일 로드
        wb = load_workbook(file_path)
        ws = wb.active

        # 빈 열 찾기
        empty_column = None
        for col in ws.iter_cols(min_row=1, max_row=1):
            if all(cell.value is None for cell in col):
                empty_column = col[0].column
                break

        # 빈 열이 없으면 새 열 추가
        if empty_column is None:
            empty_column = ws.max_column + 1

        # 헤더 추가 및 스타일 설정
        ws.cell(row=1, column=empty_column, value="전처리 후").alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

        # 응답 데이터를 기존 시트에 업데이트
        for row_index, response in enumerate(df["응답"], start=2):  # 헤더를 제외하고 2행부터 시작
            ws.cell(row=row_index, column=empty_column, value=response).alignment = Alignment(wrap_text=True)

        # 파일 저장
        wb.save(file_path)
        print("엑셀 파일이 업데이트되었습니다.")
    except Exception as e:
        print(f"엑셀 파일 업데이트 중 오류 발생: {e}")

# 메인 루프
if __name__ == "__main__":
    try:
        # Excel 파일 읽기
        df = pd.read_excel(excel_file_path)

        # "질문" 열 데이터 가져오기
        if "질문" not in df.columns:
            raise ValueError("Excel 파일에 '질문' 열이 존재하지 않습니다.")

        questions = df["질문"].fillna("").tolist()

        # 응답 열 추가
        if "응답" not in df.columns:
            df["응답"] = ""

        # 사용 가능한 CPU 코어 수의 절반 계산
        max_workers = max(1, os.cpu_count() // 2)

        # 멀티프로세싱을 사용하여 질문 처리
        with ProcessPoolExecutor(max_workers=max_workers) as executor:
            results = list(tqdm(executor.map(process_question, enumerate(questions)), total=len(questions), desc="질문 처리 중"))

        # 결과 저장
        for index, response in results:
            df.at[index, "응답"] = response

        # 기존 엑셀 파일 업데이트
        save_to_existing_excel(excel_file_path, df)

    except Exception as e:
        print(f"오류가 발생했습니다: {e}")